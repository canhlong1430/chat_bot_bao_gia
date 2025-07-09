from openpyxl import load_workbook
from openpyxl.utils import rows_from_range, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.cell import MergedCell
from copy import copy
import os, datetime
import re
from openpyxl.drawing.image import Image as OpenpyxlImage


class TemplateService:
    def __init__(self, template_folder, export_folder):
        self.template_folder = template_folder
        self.export_folder = export_folder

    def get_master(self, ws, cell):
        """
        Nếu cell là MergedCell, trả về ô top-left; ngược lại trả về chính cell.
        """
        if not isinstance(cell, MergedCell):
            return cell
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                return ws[mr.coord.split(':')[0]]
        return cell

    def safe_set(self, ws, ref, value):
        """
        Gán giá trị vào ô ref; nếu là merged, gán vào ô đầu vùng.
        """
        cell = ws[ref]
        master = self.get_master(ws, cell)
        master.value = value

    def copy_cell_format(self, ws, src_row, src_col, tgt_row, tgt_col):
        """
        Sao chép format từ ô (src_row,src_col) sang (tgt_row,tgt_col).
        """
        src = ws.cell(row=src_row, column=src_col)
        raw_tgt = ws.cell(row=tgt_row, column=tgt_col)
        tgt = self.get_master(ws, raw_tgt)
        tgt._style = copy(src._style)
        tgt.number_format = src.number_format
        tgt.font = copy(src.font)
        tgt.border = copy(src.border)
        tgt.fill = copy(src.fill)
        tgt.alignment = copy(src.alignment)
        tgt.protection = copy(src.protection)

    def delete_range(self, ws, range_str):
        """
        Xóa trắng toàn bộ nội dung, format, metadata của các hàng trong range_str,
        sau đó xoá các hàng đó.
        """
        rows_to_delete = set()
        for row in rows_from_range(range_str):
            for coord in row:
                rows_to_delete.add(ws[coord].row)
        # Lấy style mặc định để reset
        default = ws.cell(row=100, column=20)
        # Clear content và format
        for r in sorted(rows_to_delete):
            for c in range(1, 20):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    for mr in list(ws.merged_cells.ranges):
                        if cell.coordinate in mr:
                            ws.unmerge_cells(str(mr))
                self.safe_set(ws, cell.coordinate, None)
                cell._style = copy(default._style)
                cell.number_format = default.number_format
                cell.font = copy(default.font)
                cell.border = copy(default.border)
                cell.fill = copy(default.fill)
                cell.alignment = copy(default.alignment)
                cell.protection = copy(default.protection)
        # Xóa hàng
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)

    def extract_range_format(self, ws, range_str):
        """
        Trả về list thông tin format & value của từng cell trong range_str.
        Chỉ lưu value cho master cell (ô top-left của vùng merge), các ô còn lại value=None.
        """
        result = []
        for row in rows_from_range(range_str):
            for coord in row:
                cell = ws[coord]
                master = self.get_master(ws, cell)
                is_master = (cell.coordinate == master.coordinate)
                # Lưu value cho master cell, các cell còn lại lưu value thực tế (nếu không phải merged)
                if isinstance(cell, MergedCell):
                    value = master.value if is_master else None
                else:
                    value = cell.value
                result.append({
                    'coord': coord,
                    'value': value,
                    'style': copy(master._style),
                    'num_fmt': master.number_format,
                    'font': copy(master.font),
                    'border': copy(master.border),
                    'fill': copy(master.fill),
                    'align': copy(master.alignment),
                    'prot': copy(master.protection),
                    'is_master': is_master
                })
        return result

    def apply_range_format(self, ws, range_str, format_list, row_offset=0, col_offset=0):
        """
        Áp dụng list format & value vào vùng range_str, có thể dịch chuyển theo offset.
        Chỉ gán value cho cell có value khác None (theo đúng thứ tự list format).
        Thứ tự format_list phải đúng thứ tự duyệt từng cell trong rows_from_range(range_str)!
        """
        idx = 0
        for row in rows_from_range(range_str):
            for coord in row:
                fmt = format_list[idx]
                cell = ws[coord]
                tgt_row = cell.row + row_offset
                tgt_col = cell.column + col_offset
                tgt = ws.cell(row=tgt_row, column=tgt_col)
                # Chỉ gán value nếu value khác None (bám sát logic copy gốc)
                if fmt['value'] is not None:
                    self.safe_set(ws, tgt.coordinate, fmt['value'])
                tgt._style = copy(fmt['style'])
                tgt.number_format = fmt['num_fmt']
                tgt.font = copy(fmt['font'])
                tgt.border = copy(fmt['border'])
                tgt.fill = copy(fmt['fill'])
                tgt.alignment = copy(fmt['align'])
                tgt.protection = copy(fmt['prot'])
                idx += 1

    def extract_merge_list(self, ws, r_range):
        """
        Trả về list các merge cell (dưới dạng string) nằm trong vùng r_range.
        """
        area = CellRange(r_range)
        mlist = []
        for mc in ws.merged_cells:
            if mc.coord not in area:
                continue
            mlist.append(mc.coord)
        return mlist

    def apply_merge_list(self, ws, merge_list, row_offset=0, col_offset=0):
        """
        Merge lại các vùng đã lưu (dưới dạng list string), có thể dịch chuyển vị trí.
        """
        for m in merge_list:
            try:
                cr = CellRange(m)
                if row_offset or col_offset:
                    cr.shift(row_shift=row_offset, col_shift=col_offset)
                ws.merge_cells(cr.coord)
            except Exception as e:
                print(f"⚠️ WARN: Cannot merge {m}: {e}")

    import re

    def update_formula_references(self, ws, row_offset=0):
        """
        Duyệt qua tất cả cell trong ws, nếu là công thức thì cập nhật lại các tham chiếu dòng theo row_offset.
        - Với hàm SUM, chỉ cập nhật phần tham chiếu phía sau (vế phải của dấu hai chấm)
        - Với phép cộng/trừ thông thường, cập nhật cả hai vế
        """
        cell_ref_pattern = re.compile(r'([A-Z]+)(\d+)', re.IGNORECASE)
        sum_range_pattern = re.compile(r'(=\s*SUM\s*\(\s*([A-Z]+)(\d+):([A-Z]+)(\d+)\s*\))', re.IGNORECASE)

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Xử lý riêng cho hàm SUM(A1:A5)
                    def sum_repl(m):
                        # m.group(2): col1, m.group(3): row1, m.group(4): col2, m.group(5): row2
                        # Giữ nguyên row1, chỉ cộng offset cho row2
                        return f"=SUM({m.group(2)}{m.group(3)}:{m.group(4)}{int(m.group(5)) + row_offset})"

                    if sum_range_pattern.match(formula):
                        # Nếu là SUM(A1:A5), chỉ cập nhật row2
                        new_formula = sum_range_pattern.sub(sum_repl, formula)
                        cell.value = new_formula
                    else:
                        # Các công thức khác: cập nhật tất cả các tham chiếu
                        def repl(m):
                            col, rownum = m.group(1), int(m.group(2))
                            return f"{col}{rownum + row_offset}"

                        new_formula = cell_ref_pattern.sub(repl, formula)
                        cell.value = new_formula

    def export_quote(self, items):
        tpl = os.path.join(self.template_folder, 'baogia_template.xlsx')
        wb = load_workbook(tpl)
        ws = wb.active
        images = ws._images.copy() if hasattr(ws, '_images') else []

        start, end = 15, 30
        n = max(len(items) - 1, 0)
        last_col = get_column_letter(ws.max_column)
        row_range = f"A16:H24"
        row_range_to_delete = f"A16:H200"
        # Bước 1: copy + move
        if n > 0:
            extract_merge_list = self.extract_merge_list(ws, row_range)
            format_list = self.extract_range_format(ws, row_range)
            self.delete_range(ws, row_range_to_delete)
            self.apply_merge_list(ws, extract_merge_list, row_offset=n, col_offset=0)
            self.apply_range_format(ws, row_range, format_list, row_offset=n, col_offset=0)

        # Bước 2: copy format từ row15
        for r in range(start, start + n + 1):
            for c in range(1, ws.max_column + 1):
                self.copy_cell_format(ws, 15, c, r, c)

        # Bước 3: ghi data
        row = start
        for idx, item in enumerate(items):
            self.safe_set(ws, f'A{row}', idx + 1)
            self.safe_set(ws, f'C{row}', item['name'])
            self.safe_set(ws, f'D{row}', item['quantity'])
            self.safe_set(ws, f'E{row}', item['unit_price1'])
            self.safe_set(ws, f'F{row}', item['unit_price1'] * item['quantity'])
            self.safe_set(ws, f'G{row}', item['unit_price'])
            self.safe_set(ws, f'H{row}', item['unit_price'] * item['quantity'])

            extra = item.get('extra_data', {})
            row += 1

        self.update_formula_references(ws, row_offset=n)

        # === CHÈN LẠI LOGO ===
        logo_path = os.path.join(self.template_folder, 'logo.png')
        if os.path.exists(logo_path):
            logo_img = OpenpyxlImage(logo_path)
            logo_img.width = 160  # chỉnh lại nếu muốn
            logo_img.height = 139
            logo_img.anchor = 'B2'  # vị trí muốn dán
            ws.add_image(logo_img)

        os.makedirs(self.export_folder, exist_ok=True)
        out_name = f"BaoGia_{datetime.datetime.now():%Y%m%d%H%M%S}.xlsx"
        wb.save(os.path.join(self.export_folder, out_name))
        wb.close()
        return out_name
